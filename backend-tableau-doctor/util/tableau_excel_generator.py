"""
TableauExcellGenerator Class

This class is responsible for generating and formatting Excel spreadsheets 
that contain data related to Tableau projects, reports, and workbooks. 
It utilizes the pandas and openpyxl libraries to handle data and Excel file 
creation.

Attributes:
- package_list (list): A list of dictionaries, where each dictionary contains
  information about a sheet, including 'sheet_name', 'payload', and 'columns'.
- output_directory (str): The directory where the generated Excel file will 
  be saved. Default is './output'.
- file_path (str): The full path to the generated Excel file, combining the 
  output directory and the file name 'metadata.xlsx'.
- image_path (str): The path to the logo image that will be added to the 
  summary sheet.

Methods:
- __init__(package): Initializes the TableauExcellGenerator with the provided 
  package list and sets up the output directory and file paths.
  
- generate_spreadsheet(): Generates an Excel spreadsheet with multiple sheets 
  based on the provided package list. Each sheet corresponds to a project and 
  contains data from the payload. If a sheet's payload is empty, it will be 
  skipped.

- format_excel(): Applies formatting to the generated Excel spreadsheet. 
  This includes setting styles for headers, adjusting column widths based on 
  content, and enabling word wrapping for specific columns (e.g., the "Query" 
  column in the 'Custom Query Details' sheet).

- generate_summary_sheet(unique_counts): Creates or updates a summary sheet 
  in the Excel file with counts of workbooks and their types. It adds an image 
  at the top and formats the sheet with appropriate headers, borders, and 
  column widths. The counts are based on the provided unique_counts dictionary.

Raises:
- FileNotFoundError: If the specified file path cannot be found during 
  operations.
- PermissionError: If there are permission issues when accessing the file.
- ValueError: If there are issues with the values being processed.
- IndexError: If there is an attempt to access an index that does not exist 
  (e.g., when adding counts to the summary sheet).
- TypeError: If there is a type mismatch when processing the data.
- Exception: For any other unexpected runtime errors during execution.
"""

import openpyxl.workbook
import pandas as pd
import os
import logging
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.drawing.image import Image
from util.auth_clients.tableau_auth import TableauAuthClient
from util.config_managers.tableau_reader import TableauConfigManager
from openpyxl.styles import Alignment


logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

class TableauExcellGenerator:
    def __init__(self, package):
        self.package_list = package
        # Use the TableauConfigManager to read config values instead of incorrectly
        # instantiating TableauAuthClient with the class itself.
        config = TableauConfigManager()
        self.output_directory = config.get_output_directory()
        file_Timestamp = pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')
        self.file_path = f'{self.output_directory}/Tableau_Metadata_{file_Timestamp}.xlsx'
        self.image_path = config.get_logo_path()

    def generate_spreadsheet(self):
        logging.info(f'Starting to generate spreadsheet: {self.file_path}')
        try:
            # check if the output directory exists, it not create it
            os.makedirs(self.output_directory, exist_ok=True)
            
            with pd.ExcelWriter(self.file_path, engine='openpyxl') as writer:
                logging.info('Opened Excell Writer')
                
                for package in self.package_list:
                    sheet_name = package['sheet_name']
                    payload = package['payload']
                    headers = package['columns']                   
                    
                    logging.info(f'Processing sheet: {sheet_name}')
                    
                    # if not package['payload']:
                    #     logging.info(f'Skipping empty sheet: {sheet_name}')    
                    #     continue
                    # # ✅ FIX: Check for empty payload (None, empty list, etc.)
                    # if not payload or (isinstance(payload, list) and len(payload) == 0):
                    #     logging.info(f'Skipping empty sheet: {sheet_name}')    
                    #     continue
                    
                    # logging.info(f'Creating DataFrame for sheet: {sheet_name}')
                    # df = pd.DataFrame(payload)
                    
                    # logging.info(f'Writing data to sheet: {sheet_name}')
                    # df.to_excel(
                    #     writer,
                    #     sheet_name=sheet_name,
                    #     header=headers,
                    #     index=False
                    # )
                    # logging.info(f'Successfully wrote data to sheet: {sheet_name}')
                    # ✅ ENHANCED FIX: More robust empty payload detection
                    # Check if payload is None, empty list, empty dict, or falsy
                    is_empty = False
                    
                    if payload is None:
                        is_empty = True
                    elif isinstance(payload, (list, tuple)):
                        is_empty = len(payload) == 0
                    elif isinstance(payload, dict):
                        is_empty = len(payload) == 0
                    elif not payload:  # Catches other falsy values
                        is_empty = True
                    
                    if is_empty:
                        logging.info(f'Skipping empty sheet: {sheet_name} (payload is empty)')
                        continue
                    
                    logging.info(f'Creating DataFrame for sheet: {sheet_name}')
                    try:
                        df = pd.DataFrame(payload)
                        
                        # Double-check: if DataFrame is empty after creation, skip it
                        if df.empty:
                            logging.info(f'Skipping sheet: {sheet_name} (DataFrame is empty after creation)')
                            continue
                        
                        logging.info(f'Writing data to sheet: {sheet_name}')
                        df.to_excel(
                            writer,
                            sheet_name=sheet_name,
                            header=headers,
                            index=False
                        )
                        logging.info(f'Successfully wrote data to sheet: {sheet_name}')
                    except ValueError as ve:
                        # If pandas fails to create DataFrame or write it, log and skip
                        logging.warning(f'Skipping sheet {sheet_name} due to ValueError: {ve}')
                        continue
                        
        except FileNotFoundError as e:
            logging.error(f'File Not Found Error: {e}')
            raise
        except PermissionError as e:
            logging.error(f'Permission Error: {e}')
            raise
        except ValueError as e:
            logging.error(f'Value Error: {e}')
            raise
        except pd.errors.EmptyDataError as e:
            logging.error(f'Pandas: Empty Data Error: {e}')
            raise
        except pd.errors.ParserError as e:
            logging.error(f'Pandas: Parser Error: {e}')
            raise
        except Exception as e:
            logging.critical(f'Runtime Critical Error: {e}')
            raise
        
    def format_excel(self):
        logging.info(f'Starting to format spreadsheet: {self.file_path}')

        try:
            # Load the workbook and iterate over each sheet
            wb = openpyxl.load_workbook(self.file_path)

            # Define the styles
            bold_font = Font(bold=True)
            light_blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
            green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # light green
            border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                logging.info(f'Formatting sheet: {sheet_name}')

                # Set starting row for headers and data
                header_start_row = 1
                data_start_row = header_start_row + 1

                # Build header map: header name (stripped) -> column index (int)
                header_map = {}
                for cell in ws[header_start_row]:
                    header_value = str(cell.value).strip() if cell.value is not None else ""
                    if header_value:
                        header_map[header_value] = cell.col_idx  # col_idx is integer index

                # Apply styles to the header
                for cell in ws[header_start_row]:
                    cell.font = bold_font
                    cell.fill = light_blue_fill
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left', vertical='center')

                # Apply borders to the data rows and apply conditional green-fill rows where required
                for row_idx in range(data_start_row, ws.max_row + 1):
                    # Determine whether to color this row green based on rules
                    apply_green = False

                    if sheet_name == "Datasource Details" and "Used In Sheet" in header_map:
                        col_idx = header_map["Used In Sheet"]
                        value = ws.cell(row=row_idx, column=col_idx).value
                        if isinstance(value, str):
                            if value.strip().upper() == "Y":
                                apply_green = True
                        else:
                            # In case it's stored as boolean or other type, do a fallback compare
                            try:
                                if str(value).strip().upper() == "Y":
                                    apply_green = True
                            except Exception:
                                pass

                    if sheet_name == "Dashboard Details" and "Field Type" in header_map:
                        col_idx = header_map["Field Type"]
                        value = ws.cell(row=row_idx, column=col_idx).value
                        if isinstance(value, str):
                            if value.strip().lower() == "calculatedfield".lower():
                                apply_green = True
                        else:
                            try:
                                if str(value).strip().lower() == "calculatedfield".lower():
                                    apply_green = True
                            except Exception:
                                pass

                    # Apply border (and green fill if needed) to every cell in this row
                    for col_idx in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.border = border
                        if apply_green:
                            cell.fill = green_fill

                # Adjust column width based on content
                if sheet_name == 'Custom Query Details':
                    # Enable word wrapping for all columns
                    for row in ws.iter_rows(min_row=data_start_row, max_row=ws.max_row):
                        for cell in row:
                            cell.alignment = Alignment(wrap_text=True, vertical='top')
                            cell.border = border
                    
                    # Set specific widths for each column
                    ws.column_dimensions['A'].width = 20  # Project ID
                    ws.column_dimensions['B'].width = 20  # Project
                    ws.column_dimensions['C'].width = 38  # Workbook ID
                    ws.column_dimensions['D'].width = 20  # Workbook
                    ws.column_dimensions['E'].width = 38  # Custom Query ID
                    ws.column_dimensions['F'].width = 20  # Custom Query
                    ws.column_dimensions['G'].width = 100  # Query column - main content
                    
                    # Auto-adjust row heights based on content
                    for row_num in range(data_start_row, ws.max_row + 1):
                        # Calculate required height based on wrapped text
                        max_lines = 1
                        for col_num in range(1, ws.max_column + 1):
                            cell = ws.cell(row=row_num, column=col_num)
                            if cell.value:
                                # Estimate number of lines needed
                                text_length = len(str(cell.value))
                                col_width = ws.column_dimensions[cell.column_letter].width
                                estimated_lines = max(1, int(text_length / col_width) + 1)
                                max_lines = max(max_lines, estimated_lines)
                        
                        # Set row height (approximately 15 points per line)
                        ws.row_dimensions[row_num].height = max_lines * 15
                else:
                    for column in ws.columns:
                        max_length = 0
                        column_cells = [cell for cell in column]
                        for cell in column_cells:
                            try:
                                if cell.value is not None and len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except Exception:
                                pass

                        adjusted_width = max_length + 2  # Adding some extra space
                        # column_cells[0] is the header cell; use its column_letter
                        try:
                            col_letter = column_cells[0].column_letter
                            ws.column_dimensions[col_letter].width = adjusted_width
                        except Exception:
                            # fallback if something unexpected happens
                            pass

                ws.sheet_view.showGridLines = False
                ws.freeze_panes = ws.cell(row=data_start_row, column=1)

            # Save the formatted workbook
            wb.save(self.file_path)
            logging.info(f'Successfully formatted spreadsheet: {self.file_path}')

        except FileNotFoundError as e:
            logging.error(f'File Not Found Error: {e}')
            raise
        except PermissionError as e:
            logging.error(f'Permission Error: {e}')
            raise
        except Exception as e:
            logging.critical(f'Runtime Critical Error: {e}')
            raise
    
    def generate_summary_sheet(self, unique_counts, columns=None):
        try:
            wb = openpyxl.load_workbook(self.file_path)

            # Check if the 'Summary' sheet already exists
            if 'Summary' in wb.sheetnames:
                ws = wb['Summary']
                # ws.delete_rows(1, ws.max_row)  # clear old content
            else:
                ws = wb.create_sheet('Summary', 0)

            # Add the image only if it's not already added
            if ws._images == []:
                img = Image(self.image_path)
                img.width = 287
                img.height = 84
                ws.add_image(img, 'A1')

                # Hide gridlines under image (white fill)
                for row in ws.iter_rows(min_row=1, max_row=6, min_col=1, max_col=5):
                    for cell in row:
                        cell.border = None  # Ensure no border
                        cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

            # ---------- Styling ----------
            header_font = Font(bold=True, size=11)    # header text
            count_font  = Font(bold=True, size=22, color="FF7030A0")     # count/formula row
            light_blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
            # border = Border(left=Side(style='thin'), right=Side(style='thin'),
            #                 top=Side(style='thin'), bottom=Side(style='thin'))
            # Border styles
            medium_border = Side(style='medium')  # thicker for visible outline
            no_border = Side(style=None)

            center_alignment = Alignment(horizontal="center", vertical="center")

            # ---------- Header Row ----------
            top_header_row = 7
            headers = [
                "Total no of Dashboards",
                "Total no of Reports",
                "Total no of Report Fields",
                "Total no of Datasources",
                "Total no of DB Tables",
                "Total no of DB Columns",
                "Total no of Custom SQLs"
            ]

            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=top_header_row, column=col, value=header)
                cell.font = header_font
                cell.fill = light_blue_fill
                cell.alignment = center_alignment
                #cell.border = border

                # Thick verticals + top border; no bottom border (merges visually with count row)
                cell.border = Border(
                    left=medium_border,
                    right=medium_border,
                    top=medium_border,
                    bottom=no_border
                )
                ws.column_dimensions[cell.column_letter].width = len(header)
                
            # ---------- Formula Row ----------
            header_row = 11
            formula_row = top_header_row + 1
            # ---------- Formula Row (Dynamic) ----------
            data_start_row = header_row + 1
            data_end_row = header_row + len(unique_counts)
            formula_columns = ['B', 'C', 'D', 'G', 'H', 'I', 'J']
            formulas = [
                f"=SUM({col}{data_start_row}:{col}{data_end_row})"
                for col in formula_columns
            ]

            
            center_alignment = Alignment(horizontal="center", vertical="center")
            for col, formula in enumerate(formulas, start=1):
                cell = ws.cell(row=formula_row, column=col, value=formula)
                cell.font = count_font
                cell.alignment = center_alignment  # centered horizontally & vertically
                cell.fill = light_blue_fill
                # Thick verticals + bottom border; no top border (merges visually with header)
                cell.border = Border(
                    left=medium_border,
                    right=medium_border,
                    top=no_border,
                    bottom=medium_border
                )
 
            # Define styles
            bold_font = Font(bold=True)
            light_blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
            border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))
 
            # Determine columns from unique_counts (fall back to empty list if none)
            if unique_counts and isinstance(unique_counts, list) and len(unique_counts) > 0:
                columns = list(unique_counts[0].keys())
            else:
                columns = []
 
            for col_idx, col_name in enumerate(columns, start=1):
                cell = ws.cell(row=header_row, column=col_idx, value=col_name)
                cell.font = bold_font
                cell.fill = light_blue_fill
                cell.border = border
 
            # ✅ 4. Write data rows
            for row_idx, entry in enumerate(unique_counts, start=header_row + 1):
                for col_idx, col_name in enumerate(columns, start=1):
                    value = entry.get(col_name, "")
                    c = ws.cell(row=row_idx, column=col_idx, value=value)
                    c.border = border
 
            # ✅ 5. Auto-fit column width
            for column in ws.columns:
                max_len = max((len(str(cell.value)) if cell.value else 0) for cell in column)
                ws.column_dimensions[column[0].column_letter].width = max_len

            # ---------- Visual Formatting ----------
            # Match row heights to blend perfectly
            #ws.row_dimensions[header_row].height = 25
            #ws.row_dimensions[formula_row].height = 40
            ws.sheet_view.showGridLines = False
            #ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
            ws.freeze_panes = ws.cell(row=top_header_row + 5, column=1)

            # ---------- Save ----------
            wb.save(self.file_path)
            logging.info(f"Successfully created the formatted 'Summary' sheet: {self.file_path}")

        except FileNotFoundError as e:
            logging.error(f'File Not Found Error: {e}')
            raise
        except ValueError as e:
            logging.error(f'Value Error: {e}')
            raise
        except IndexError as e:
            logging.error(f'Index Error: {e}')
            raise
        except TypeError as e:
            logging.error(f'Type Error: {e}')
            raise
        except PermissionError as e:
            logging.error(f'Permission Error: {e}')
            raise
        except Exception as e:
            logging.critical(f'Runtime Critical Error: {e}')
            raise